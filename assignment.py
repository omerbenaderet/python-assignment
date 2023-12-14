from openpyxl import load_workbook, Workbook
import re
from datetime import datetime
# Load Logs Workbook
data_file = './Logs - SRE_Techops Candidate Assignment.xlsx'
logwb = load_workbook(data_file)
logws = logwb['Logs.csv']
logs = list(logws.rows)

# Strip with Regex to fields lists
timestamp = []
service = []
orgs = []
status = []
dur = []
# Additional list for seconds only value
dursec = []
for log in logs[1:]:
    match = re.match(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*? [A-Z]+ (\w+) (\w+) .*? \((\w+)\), took (.*?) \((\d+).*', str(log[0].value))
    timestamp.append(match.group(1))
    service.append(match.group(2))
    orgs.append(match.group(3))
    status.append(match.group(4))
    dur.append(match.group(5))
    dursec.append(match.group(6))
alldata = [timestamp, service, orgs, status, dur]

# Create new workbook
wb = Workbook()
ws = wb.active

# Add titles
ws["A1"] = "Timestamp"
ws["B1"] = "Service"
ws["C1"] = "Organization"
ws["D1"] = "Status"
ws["E1"] = "Duration"

# Function for adding new data to cell
def addcell(row, column, data):
    ws.cell(row=row, column=column, value=data)

# Adding data to the new file
row = 2
column = 1
for field in alldata:
    for i in field:
        addcell(row, column, i)
        row += 1
    column += 1
    row = 2

# Save file
wb.save('./Parsed-Logs.xlsx')

# Function to initialize average dictionaries 
def initialize_dicts(org, srv, ts, duration_avg, between_avg, i):
    ts_datetime = datetime.strptime(ts[i], "%Y-%m-%d %H:%M:%S")
    duration_avg[(org[i], srv[i])] = [0, 0, 0]
    between_avg[(org[i], srv[i])] = [ts_datetime, 0, 0, 0]
    return duration_avg, between_avg

# Function to determine average time between successful runs 
def update_between_avg(org, srv, ts, between_avg, i):
    ts_datetime = datetime.strptime(ts[i], "%Y-%m-%d %H:%M:%S")
    diff = between_avg[(org[i], srv[i])][0] - ts_datetime
    between_avg[(org[i], srv[i])][0] = ts_datetime
    between_avg[(org[i], srv[i])][1] += diff.total_seconds()
    between_avg[(org[i], srv[i])][2] += 1
    between_avg[(org[i], srv[i])][3] = between_avg[(org[i], srv[i])][1] / between_avg[(org[i], srv[i])][2]

# Function to calculate average duration of successful runs
def update_duration_avg(org, srv, ts, duration_avg, i):
    duration_avg[(org[i], srv[i])][0] += int(dursec[i])
    duration_avg[(org[i], srv[i])][1] += 1
    duration_avg[(org[i], srv[i])][2] = duration_avg[(org[i], srv[i])][0] / duration_avg[(org[i], srv[i])][1]

# Function to convert seconds to hours, minutes, seconds for printing
def seconds_to_hms(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    result = ""
    if hours > 0:
        result += f"{int(hours)} {'hour' if int(hours) == 1 else 'hours'}"
    if minutes > 0:
        if result:
            result += ", "
        result += f"{int(minutes)} {'minute' if (minutes) == 1 else 'minutes'}"
    if seconds > 0 or not result:
        if result:
            result += ", "
        result += f"{int(seconds)} {'second' if int(seconds) == 1 else 'seconds'}"
    return result

# Function to print averages per organization-service pair
def print_avg(org, srv, duration_avg, between_avg):
    for i in range(0, len(duration_avg)):
        # If case there is only one successful run
        if between_avg[(org[i], srv[i])][3] == 0:
            between_avg[(org[i], srv[i])][3] = duration_avg[(org[i], srv[i])][2]
        print(f"For Organization " + str(list(duration_avg)[i][0]) + " And Service " + str(list(duration_avg)[i][1]) + ":\n" \
              " The average duration of successful runs is: " + seconds_to_hms(duration_avg[(org[i], srv[i])][2]) + "\n" \
              " The average time between successful runs is: " + seconds_to_hms(between_avg[(org[i], srv[i])][3]))

# Function for calculating average duration of successful runs and average time between them
def DataAnalysis(org, srv, ts):
# duration_avg = {(org, srv) : [sum, count, avg]} - dictionary for average duration of successful runs
# between_avg = {(org, srv) : [lasttimestamp, sum, count, avg]} - dictionary for average time between successful runs
    duration_avg = {}
    between_avg = {}
    for i in range(0, len(org)):
        # Only if the status of the log is successfull
        if status[i] == "SUCCESS":
            if (org[i], srv[i]) not in duration_avg:
                initialize_dicts(org, srv, ts, duration_avg, between_avg, i)
            else:
                update_between_avg(org, srv, ts, between_avg, i)
            update_duration_avg(org, srv, ts, duration_avg, i)
    print_avg(org, srv, duration_avg, between_avg)

DataAnalysis(orgs, service, timestamp)